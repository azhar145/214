     
def t():
    import os
    import subprocess
    import shutil
##    cmd='copy C://Users//aa300j//Downloads//info2.txt C://Users//aa300j//Downloads//p34'
##    status = subprocess.call(cmd, shell=True)


    path=r'C:\Users\aa300j\Downloads\p34'
    p=os.listdir(path)
    
    print(p)
    t=str('copy ')+ str(path)+str('\\')+str('instance_relationshipTenantSIInstanceProdZoneATN3_2.6.csv')+ str('  ')+str(path)+str('\\')
    
##    os.system(t)
##    os.system("copy C://Users//aa300j//Downloads//p34//instance_relationshipTenantSIInstanceProdZoneATN3_2.6.csv C://Users//aa300j//Downloads//p34//mm.csv")
    p = os.popen(str('copy C:\\Users\\aa300j\\Downloads\\p34\\instance_relationshipTenantSIInstanceProdZoneATN3_2.6.csv C:\\Users\\aa300j\\Downloads\\p34\\m')).read() 
##    subprocess.run(t)
    print(p)
    print('\n\n')
    print(p)
    












def change_filenames_for_bulk_import(version_no,g3):
    import pandas as pd
    import os,numpy
    import sys
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    import warnings
    import os
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    from openpyxl.styles import Font, Fill
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    import os.path
    import os
    import subprocess
    import shutil
    import csv
    import fnmatch
    import streamlit as st

    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)


#######################
    path = os.path.abspath(__file__)
    dir_path = (path)
    pp = str('Curr_dir_Curr_Module --> ')+str(dir_path)+str('   ')+str('in module -----> [change_filenames_for_bulk_import]')
    print('\n')
    print('##################################################')
    print('***** where am i?   ')
    print(pp)
    print('******************** 334')
    print('##################################################')
    print('\n')
    print('\n')

################
    print('--- inputs')
    print('1) version_no: ',version_no)
    print('2) g3 :',g3)
    print('\n\n')
    


    #####################################################################################
    # input fields:
##    t=r'C:\Users\aa300j\Downloads\New folder (3)\ff'
##    t=r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4_delete55'
##    t2=r'C:\Users\aa300j\Downloads\Candidate_DLP_files_x4_delete55'
    t=g3
    t2=g3
    input_directory=str(t)

    output_directory=str(t2)+str('\\')+str('bulk')
    if not os.path.isdir(output_directory):
        os.mkdir(output_directory)
    ends_with='.csv'
    split_by=' '
    what_to_change='__bbbb___'
    version_no=version_no
    import pandas as pd
    

    #####################################################################################
    k=0
    c3=0

    print(os.listdir(input_directory),' 0000000000000000000000000000000000')


##    st.header("Counter")
##
##    if 'counter' not in st.session_state:
##        st.session_state.counter = 0
##
##    button = st.button('Increment')
##
##    if button:
##        st.session_state.counter += 1
##
##    st.write('Counter = ', st.session_state.counter)
    
    for f in os.listdir(input_directory):
##        print(f)
        if '.csv' in f:
            df=pd.read_csv(str(input_directory)+str('\\')+str(f))
            f2=str(f).split(' ')
    ##        print(f2,'   ',input_directory)
    ##        print(len(f2))
##            if len(f2) > 1:
##                m=str(f2).split(' ')[1]

            print('==========   ',f, ' 4444 azhar')
##            placeholder = st.empty()
##            placeholder.write(f)
##            placeholder = st.empty()
##            st.header(f)

            st.markdown(f'<p small style="color:#333dff;font-size:10px;margin:0;padding:0;line-height:0px;">{f}</small>', unsafe_allow_html=True)
##            st.text(f)
##            st.code_util(f)
            
##            st.write('<p style="font-size:16px; color:red;">f</p>',\
##                     unsafe_allow_html=True) 
            

            m=f
##            print(m)

            
##            m2=str(m).split('.csv')[0]
##            m2=str(m2).replace("'","")
            m2=str(m).split(' ')[1]
            m2=str(m2).split('.csv')[0]
            
##            m2=str(m2).replace("'","")
    ##        print('m2= ',m2)
    ##
            g=str('instance~')+str(m2)+str('~')+str(version_no)+str('.csv')

    ##        pathv=str(output_directory)+str('\\')+str(g)
    ##        os.rename(src, dest)
    ##        write.csv(os.path.join(output_directory, g))
            df.to_csv(str(output_directory)+str('\\')+str(g),index=False)


            
            k=k+1
        c3=c3+1
##        f.to_csv(fullname)
##        print(g)
    
    z5=['vlanCharacteristicsInstanceProdRegion','ProdZone']    
    for f in os.listdir(str(g3)+str('\\')+str('bulk')):
        path=os.path.join(str(g3)+str('\\')+str('bulk'),f)

        for p in z5:
            if p in f:
                os.remove(path)
                print('\n\n')
                if 'vlanCharacteristicsInstanceProdRegion' in p:
                    print('********************** vlan removed', 'code 668')
                if 'ProdZone' in p:
                    print('********************** ProdZone', 'code 668')
                print('\n\n')    


        
##
    print('\n\n')            
    g4_count = len(fnmatch.filter(os.listdir(str(g3)+str('\\')+str('bulk')), '*.*'))
    print('1) Bulk tool finished - file_names converted and saved in bulk folder no of files= ',k,'    ',str(g3)+str('\\')+str('bulk'))
    print('2) Bulk tool finished - [3 files removed vlanchar,zone] ',g4_count,'                          ',str(g3)+str('\\')+str('bulk'))
    print('3) Missing files path -----> ',g3)

    #######################
##    path = os.path.abspath(__file__)
##    dir_path = (path)
##    pp = str('Curr_dir_Curr_Module --> ')+str(dir_path)+str('   ')+str('in module -----> [change_filenames_for_bulk_import]')
##    print('\n')
##    print('##################################################')
##    print('***** where am i?   ')
##    print(pp)
##    print('******************** 334')
##    print('##################################################')
##    print('\n')
##    print('\n')



def remove_3_files(g3):
    import pandas as pd
    import os,numpy
    import sys
    from openpyxl.styles import PatternFill
    from openpyxl.styles import Font
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    import warnings
    import os
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.workbook import Workbook
    from openpyxl.styles import Font, Fill
    from openpyxl.utils import get_column_letter
    from openpyxl.styles import Alignment

    pd.set_option('display.max_rows', None)
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', None)

    print('nnnnnnnnnnnnnnnn')

##    bulk_folder_path=r'C:\Users\aa300j\Downloads\PEP\Connect_it_files\Data_Loading\Data Management\create_DLP_files\DLP_NCX_template_(open In-Service file)\C2s_December_1_2022\HST5G\Candidate'
##    output_from_DLP_HTML_csv_file_loaded_in_DLP_sucessfully=r'C:\Users\aa300j\Desktop\Data Dictionary and Lookup Platform_HST5g_uploaded.csv'

    
    g3=r'C:\Users\aa300j\Downloads\missing_dlp'
    g4=r'C:\Users\aa300j\Downloads\3files_removed_dlp'

    pp=g3
##    pp=os.listdir(g3)
    print(str(pp))
    k=0

    gg=[]
    print('\n\n')
    tt=[]
    z5=['vlanCharacteristicsInstanceProdRegion','ProdZone']
    for x in (pp):
        
##        if 'vlanCharacteristicsInstanceProdRegion' in x or 'ProdZone' in x:
##            print(x)
        if x not in z5:
            

            tt=str('copy')+str(' ')+str(g3)+str('\\')+str(x) + str(' ')+str(g4)
            print(tt)
            p = os.popen(tt).read()
            print(p)

        if x in z5:
            tt.append(x)




            
    print('\n\n')


    for x in (pp):
        
        if 'vlanCharacteristicsInstanceProdRegion' in x or 'ProdZone' in x:
            print(x)        

    print('\n\n\n')
    print(str(pp))
    print(' # of files [minus 3 files] :',len(tt))
